"""営業用 Excel 価格シート生成器 (ブランド対応版)。

構成 (CONFIGS):
  workbridge -> SAAS_PRICING_JP.xlsx       (WorkBridge Japan, 旧価格)
  linguasync -> PRICING_LINGUASYNC_JP.xlsx (LinguaSync, 新価格 + MVP開発費)

シート:
  1. トップ        — 表紙
  2. 見積シミュレーション — 店舗数 · プラン入力で自動計算
  3. プラン比較    — 3 ティア静的比較
  4. 5年TCO比較    — 自社開発 vs SaaS 5 年総費比較
  5. オプション    — アドオン価格表
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


# ============================================================
# ブランド設定 — ここを変えるだけで再生成可能
# ============================================================

CONFIGS = {
    "workbridge": {
        "service_name": "WorkBridge Japan",
        "output_filename": "SAAS_PRICING_JP.xlsx",
        "starter_price": 12800,
        "business_price": 9800,
        "enterprise_price": 6800,
        "hq_license_default": 100000,
        # MVP 開発費を表示しない場合 (旧モデル)
        "use_mvp_fee": False,
        "mvp_dev_fee": 0,
        # 旧モデルの初期費用 (使われない場合は 0)
        "init_starter": 150000,
        "init_business": 300000,
        "init_enterprise": 800000,
        # TCO 比較 — SaaS 5 年初期費 (旧モデルでは init_enterprise を使う)
        "tco_saas_initial": 800000,
    },
    "linguasync": {
        "service_name": "LinguaSync",
        "output_filename": "PRICING_LINGUASYNC_JP.xlsx",
        "starter_price": 15800,
        "business_price": 12800,
        "enterprise_price": 9800,
        "hq_license_default": 100000,
        # MVP 開発費を表示する (新モデル)
        "use_mvp_fee": True,
        "mvp_dev_fee": 5000000,
        # 新モデルでは MVP 開発費に統合され、別途の初期費用は基本 0
        "init_starter": 0,
        "init_business": 0,
        "init_enterprise": 0,
        # TCO 比較 — SaaS 5 年初期費 = MVP 開発費
        "tco_saas_initial": 5000000,
    },
}


# ============================================================
# 共通スタイル
# ============================================================

YEN_FMT = '"¥"#,##0;[Red]-"¥"#,##0'
INT_FMT = "#,##0"
PCT_FMT = "0%"

THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK = Side(style="medium", color="333333")
BORDER_HEADER = Border(left=THIN, right=THIN, top=THICK, bottom=THICK)

H1 = Font(name="Yu Gothic", size=18, bold=True, color="1F2937")
H2 = Font(name="Yu Gothic", size=13, bold=True, color="1F2937")
H3 = Font(name="Yu Gothic", size=11, bold=True, color="374151")
BODY = Font(name="Yu Gothic", size=10, color="111827")
BODY_BOLD = Font(name="Yu Gothic", size=10, bold=True, color="111827")
SMALL = Font(name="Yu Gothic", size=9, color="6B7280")
INPUT_FONT = Font(name="Yu Gothic", size=11, bold=True, color="1E3A8A")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

FILL_HEAD = PatternFill("solid", fgColor="312E81")
FILL_SUBHEAD = PatternFill("solid", fgColor="E0E7FF")
FILL_INPUT = PatternFill("solid", fgColor="FEF3C7")
FILL_OUTPUT = PatternFill("solid", fgColor="D1FAE5")
FILL_DANGER = PatternFill("solid", fgColor="FEE2E2")
FILL_ALT = PatternFill("solid", fgColor="F9FAFB")
HEAD_FONT = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _heading(ws, row, col, text, font=H1):
    ws.cell(row=row, column=col, value=text).font = font


def _cell(ws, row, col, value, *, font=BODY, fill=None, align=None, fmt=None, border=BORDER_ALL):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill is not None:
        c.fill = fill
    if align is not None:
        c.alignment = align
    if fmt is not None:
        c.number_format = fmt
    if border is not None:
        c.border = border
    return c


# ============================================================
# 1. トップ (cover)
# ============================================================

def build_cover(wb: Workbook, cfg: dict) -> None:
    ws = wb.active
    ws.title = "トップ"
    _set_widths(ws, [4, 30, 30, 30, 30, 4])

    ws["B2"] = cfg["service_name"]
    ws["B2"].font = Font(name="Yu Gothic", size=24, bold=True, color="312E81")
    ws.merge_cells("B2:E2")

    ws["B3"] = "料金見積・ROI シミュレーション (税抜)"
    ws["B3"].font = Font(name="Yu Gothic", size=13, color="4B5563")
    ws.merge_cells("B3:E3")

    ws["B5"] = (
        "本ファイルは営業・お客様への料金提示・社内見積検討に使用するワーク\n"
        "シートです。「見積シミュレーション」シートに店舗数・プラン等を入力\n"
        "すると、自動的に月額・年額・5 年総コストおよび自社開発との比較が\n"
        "計算されます。"
    )
    ws["B5"].font = BODY
    ws["B5"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells("B5:E8")
    for r in (5, 6, 7, 8):
        ws.row_dimensions[r].height = 18

    ws["B10"] = "シート構成"
    ws["B10"].font = H2

    sheets_desc = [
        ("見積シミュレーション", "店舗数・プラン入力で月額/年額/5 年総コストを自動計算"),
        ("プラン比較", "Starter / Business / Enterprise の機能・SLA・サポート比較表"),
        ("5年TCO比較", "自社開発 (中堅 IT / 大手 SIer) vs SaaS の 5 年総額比較"),
        ("オプション", "追加で発生する可能性のある費用項目とアドオン料金"),
    ]
    for i, (name, desc) in enumerate(sheets_desc):
        r = 11 + i
        _cell(ws, r, 2, name, font=BODY_BOLD, fill=FILL_SUBHEAD, align=LEFT)
        _cell(ws, r, 3, desc, font=BODY, align=LEFT)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)

    ws["B16"] = "ご使用にあたって"
    ws["B16"].font = H2

    notes = [
        "• 黄色セル (FEF3C7) は入力項目です。値を変更すると下記が自動更新されます。",
        "• 緑色セル (D1FAE5) は自動計算結果です。直接編集しないでください。",
        "• 表記は全て税抜です。実際の見積もりには税額・送金手数料等を別途加算してください。",
        "• 想定為替: ¥150/$。為替変動が大きい場合、「5年TCO比較」シートの定義値を調整してください。",
        "• Google Translate API 単価: $20/100万字 (2026年時点)。GCP 値上げに応じて調整可能。",
    ]
    for i, line in enumerate(notes):
        ws.cell(row=17 + i, column=2, value=line).font = SMALL
        ws.merge_cells(start_row=17 + i, start_column=2, end_row=17 + i, end_column=5)

    ws["B24"] = "提案先："
    ws["B24"].font = BODY_BOLD
    _cell(ws, 24, 3, "[貴社名]", font=INPUT_FONT, fill=FILL_INPUT, align=LEFT)
    ws.merge_cells("C24:E24")

    ws["B25"] = "提案者："
    ws["B25"].font = BODY_BOLD
    _cell(ws, 25, 3, "[当社名]", font=INPUT_FONT, fill=FILL_INPUT, align=LEFT)
    ws.merge_cells("C25:E25")

    ws["B26"] = "作成日："
    ws["B26"].font = BODY_BOLD
    _cell(ws, 26, 3, "[YYYY年MM月DD日]", font=INPUT_FONT, fill=FILL_INPUT, align=LEFT)
    ws.merge_cells("C26:E26")


# ============================================================
# 2. 見積シミュレーション (calculator)
# ============================================================

def build_calculator(wb: Workbook, cfg: dict) -> None:
    ws = wb.create_sheet("見積シミュレーション")
    _set_widths(ws, [4, 28, 20, 18, 18, 18, 18, 4])

    _heading(ws, 2, 2, "見積シミュレーション", font=H1)
    _cell(ws, 3, 2, "黄色セルに値を入力すると、緑色セルが自動計算されます。",
          font=SMALL, align=LEFT, border=None)
    ws.merge_cells("B3:G3")

    _heading(ws, 5, 2, "入力項目", font=H2)

    use_mvp = cfg["use_mvp_fee"]

    # 入力項目を動的に構成
    inputs: list[tuple[str, object, str, str]] = [
        ("店舗数", 10, "INT_FMT", "ご利用予定の店舗数"),
        ("選択プラン", "Business", "TEXT", "Starter / Business / Enterprise"),
        ("Starter 月額/店舗", cfg["starter_price"], "YEN_FMT", "1 店舗のみ"),
        ("Business 月額/店舗", cfg["business_price"], "YEN_FMT", "1〜5 店舗向け"),
        ("Enterprise 月額/店舗", cfg["enterprise_price"], "YEN_FMT", "10 店舗以上 + 本部ライセンス"),
        ("Enterprise 本部ライセンス/月", cfg["hq_license_default"], "YEN_FMT", "Enterprise のみ"),
    ]
    if use_mvp:
        inputs.append(("MVP 開発費 (一括)", cfg["mvp_dev_fee"], "YEN_FMT", "全プラン共通の初期構築費"))
    else:
        inputs.extend([
            ("初期費用 (Starter)", cfg["init_starter"], "YEN_FMT", "一括"),
            ("初期費用 (Business)", cfg["init_business"], "YEN_FMT", "一括"),
            ("初期費用 (Enterprise)", cfg["init_enterprise"], "YEN_FMT", "規模により協議"),
        ])
    inputs.extend([
        ("年間契約割引 (%)", 0.10, "PCT_FMT", "10〜20% (Enterprise 年契約時)"),
        ("契約年数", 1, "INT_FMT", "ROI 計算で 1〜5 年"),
    ])

    fmt_map = {"INT_FMT": INT_FMT, "YEN_FMT": YEN_FMT, "PCT_FMT": PCT_FMT, "TEXT": "@"}

    # 各入力の行番号を覚えておく (式組立てに使う)
    rows_by_label: dict[str, int] = {}
    for i, (label, val, fmt_key, hint) in enumerate(inputs):
        r = 6 + i
        rows_by_label[label] = r
        _cell(ws, r, 2, label, font=BODY_BOLD, fill=FILL_SUBHEAD, align=LEFT)
        _cell(ws, r, 3, val, font=INPUT_FONT, fill=FILL_INPUT,
              align=RIGHT if fmt_key != "TEXT" else CENTER, fmt=fmt_map[fmt_key])
        _cell(ws, r, 4, hint, font=SMALL, align=LEFT, border=None)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)

    # セル参照 (Cn) を取得するヘルパー
    def C(label: str) -> str:
        return f"$C${rows_by_label[label]}"

    out_start = 6 + len(inputs) + 2
    _heading(ws, out_start, 2, "自動計算結果", font=H2)

    # プランごとの月額単価選択
    pick_unit = (
        f'=IF({C("選択プラン")}="Starter",{C("Starter 月額/店舗")},'
        f'IF({C("選択プラン")}="Business",{C("Business 月額/店舗")},'
        f'IF({C("選択プラン")}="Enterprise",{C("Enterprise 月額/店舗")},{C("Business 月額/店舗")})))'
    )
    pick_hq = f'=IF({C("選択プラン")}="Enterprise",{C("Enterprise 本部ライセンス/月")},0)'

    if use_mvp:
        # 新モデル: MVP 開発費は全プラン同額
        pick_initial = f'={C("MVP 開発費 (一括)")}'
        initial_label = "MVP 開発費 (一括)"
    else:
        # 旧モデル: プラン別の初期費用
        pick_initial = (
            f'=IF({C("選択プラン")}="Starter",{C("初期費用 (Starter)")},'
            f'IF({C("選択プラン")}="Business",{C("初期費用 (Business)")},'
            f'IF({C("選択プラン")}="Enterprise",{C("初期費用 (Enterprise)")},{C("初期費用 (Business)")})))'
        )
        initial_label = "初期費用 (一括)"

    # 出力行を順次配置
    out_rows: list[tuple[str, str, bool]] = [
        ("選択プラン単価/店舗/月", pick_unit, False),
        ("店舗数 × 単価/月 (税抜)", f"=$C${rows_by_label['店舗数']}*$C${out_start + 1}", False),
        ("本部ライセンス/月 (Enterprise のみ)", pick_hq, False),
    ]
    # 月額合計 = 店舗計 + 本部
    monthly_total_row = out_start + 4  # = 店舗計行(out_start+2) + 本部(out_start+3) → out_start+4
    out_rows.append(("月額合計 (税抜)", f"=$C${out_start + 2}+$C${out_start + 3}", True))
    out_rows.append(("年間契約割引額/月", f"=$C${out_start + 4}*{C('年間契約割引 (%)')}", False))
    out_rows.append(("月額 (年間契約割引適用後)", f"=$C${out_start + 4}-$C${out_start + 5}", True))
    out_rows.append(("年額 (税抜・割引後)", f"=$C${out_start + 6}*12", True))
    out_rows.append((initial_label, pick_initial, False))
    # 契約年数の総額
    out_rows.append((
        "契約年数の総額 (税抜)",
        f"=$C${out_start + 7}*{C('契約年数')}+$C${out_start + 8}",
        True,
    ))

    for i, (label, formula, highlight) in enumerate(out_rows):
        r = out_start + 1 + i
        _cell(ws, r, 2, label, font=BODY_BOLD, fill=FILL_SUBHEAD, align=LEFT)
        fill = FILL_OUTPUT if highlight else None
        font_used = BODY_BOLD if highlight else BODY
        _cell(ws, r, 3, formula, font=font_used, fill=fill, align=RIGHT, fmt=YEN_FMT)

    # 主要指標を太字緑で強調
    important_rows = [out_start + 4, out_start + 6, out_start + 7, out_start + 9]
    for rr in important_rows:
        ws.cell(row=rr, column=3).font = Font(name="Yu Gothic", size=12, bold=True, color="047857")

    note_row = out_start + len(out_rows) + 2
    ws.cell(row=note_row, column=2,
            value="※ 上記は基準価格の自動見積もりです。実際の最終価格は規模・要件・契約条件に応じ個別協議となります。"
            ).font = SMALL
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=7)


# ============================================================
# 3. プラン比較 (static)
# ============================================================

def build_plan_comparison(wb: Workbook, cfg: dict) -> None:
    ws = wb.create_sheet("プラン比較")
    _set_widths(ws, [4, 32, 24, 24, 28, 4])

    _heading(ws, 2, 2, "プラン別 機能・サービス比較", font=H1)

    yen = lambda n: f"¥{n:,}"
    rows: list[tuple[str, ...]] = [
        ("項目", "Starter", "Business", "Enterprise"),
    ]
    if cfg["use_mvp_fee"]:
        rows.append(("MVP 開発費 (一括)",
                     yen(cfg["mvp_dev_fee"]),
                     yen(cfg["mvp_dev_fee"]),
                     f"{yen(cfg['mvp_dev_fee'])} (規模により協議)"))
    rows.extend([
        ("月額/店舗 (税抜)",
         yen(cfg["starter_price"]),
         yen(cfg["business_price"]),
         f"{yen(cfg['enterprise_price'])}〜"),
        ("本部ライセンス/月", "─", "─", "¥80,000〜¥150,000"),
    ])
    if not cfg["use_mvp_fee"]:
        rows.append(("初期費用",
                     yen(cfg["init_starter"]),
                     yen(cfg["init_business"]),
                     f"{yen(cfg['init_enterprise'])} (規模協議)"))
    rows.extend([
        ("最低契約期間", "3 ヶ月", "6 ヶ月", "12 ヶ月 (年間契約割引 10〜20%)"),
        ("対象店舗数", "単独店舗のみ", "1〜5 店舗", "10 店舗以上"),
        ("スタッフ登録/店舗", "〜10 名", "〜20 名", "無制限"),
        ("月間翻訳 API 字数", "50 万字", "200 万字/店舗", "Fair Use 無制限"),
        ("超過課金", "¥1,000/10 万字", "¥800/10 万字", "─"),
        ("リアルタイム指示送受信", "✓", "✓", "✓"),
        ("6 言語自動翻訳", "✓", "✓", "✓"),
        ("やさしい日本語変換", "✓", "✓", "✓"),
        ("3 ボタン応答 + カスタム音声応答", "✓", "✓", "✓"),
        ("画像添付指示", "✓", "✓", "✓"),
        ("CSV 一括登録", "✓", "✓", "✓"),
        ("60 日分の履歴保持", "✓", "✓", "選択により最大 12 ヶ月"),
        ("PWA (iOS/Android 対応)", "✓", "✓", "✓"),
        ("店舗専用業務用語辞書", "─", "✓", "✓"),
        ("表現・言い換え辞書", "─", "✓", "✓"),
        ("店舗別ロゴ・ブランド統合", "─", "オプション", "標準"),
        ("本部一括ダッシュボード", "─", "─", "✓"),
        ("複数店舗一斉指示", "─", "─", "✓"),
        ("店舗別 売上・稼働分析", "─", "─", "✓"),
        ("SSO (SAML/OIDC)", "─", "─", "✓"),
        ("監査ログ", "─", "─", "✓"),
        ("DB 行レベル分離 (Postgres RLS)", "─", "─", "✓"),
        ("メールサポート", "✓", "✓", "✓"),
        ("チャットサポート", "─", "✓", "✓"),
        ("電話サポート", "─", "─", "✓"),
        ("専任カスタマーサクセス", "─", "─", "✓"),
        ("応答時間 (営業時間内)", "4 時間", "2 時間", "1 時間 (重大障害 24h)"),
        ("稼働率目標 (SLA)", "ベストエフォート", "99.5%", "99.9% + 障害時クレジット"),
        ("バックアップ", "日次/7 日", "日次/14 日", "日次/30 日 + オンデマンド"),
        ("オンライントレーニング", "1 回 (1 時間)", "2 回 (各 1 時間)", "無制限"),
        ("オンサイトトレーニング", "別途有料", "別途有料", "1 日含む"),
        ("ペネトレーションテスト", "─", "─", "年 1 回含む"),
        ("データ削除 SLA", "5 営業日", "5 営業日", "2 営業日"),
        ("DPA / 個人情報チェックシート", "標準", "標準 + カスタム項目", "個別協議 + DPIA"),
        ("ホスティング", "日本 (東京)", "日本 (東京)", "日本 (東京) + 専用環境オプション"),
    ])

    start_row = 4
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            r, c = start_row + i, 2 + j
            if i == 0:
                _cell(ws, r, c, val, font=HEAD_FONT, fill=FILL_HEAD, align=CENTER)
            else:
                fill = FILL_ALT if i % 2 == 1 else None
                font = BODY_BOLD if j == 0 else BODY
                align = LEFT if j == 0 else CENTER
                _cell(ws, r, c, val, font=font, fill=fill, align=align)
        ws.row_dimensions[start_row + i].height = 22

    ws.row_dimensions[start_row].height = 28
    ws.freeze_panes = ws.cell(row=start_row + 1, column=3)


# ============================================================
# 4. 5 年 TCO 比較
# ============================================================

def build_tco(wb: Workbook, cfg: dict) -> None:
    ws = wb.create_sheet("5年TCO比較")
    _set_widths(ws, [4, 32, 22, 22, 22, 4])

    _heading(ws, 2, 2, "自社開発 vs SaaS — 5 年総コスト比較", font=H1)

    _cell(ws, 4, 2, "前提: 50 店舗チェーン、5 年運用", font=BODY_BOLD, align=LEFT, border=None)
    ws.merge_cells("B4:E4")

    _heading(ws, 6, 2, "前提条件 (調整可)", font=H2)

    saas_init_label = "SaaS 初期費 (MVP 開発費 ¥)" if cfg["use_mvp_fee"] else "SaaS 初期費 (¥)"
    presets: list[tuple[str, int, str]] = [
        ("店舗数", 50, INT_FMT),
        ("運用年数", 5, INT_FMT),
        ("中堅 IT 開発初期費 (¥)", 55_000_000, YEN_FMT),
        ("中堅 IT 年間運用費 (¥)", 6_000_000, YEN_FMT),
        ("大手 SIer 開発初期費 (¥)", 150_000_000, YEN_FMT),
        ("大手 SIer 年間運用費 (¥)", 15_000_000, YEN_FMT),
        (f"{cfg['service_name']} Enterprise 月額/店舗 (¥)", cfg["enterprise_price"], YEN_FMT),
        (f"{cfg['service_name']} Enterprise 本部/月 (¥)", cfg["hq_license_default"], YEN_FMT),
        (saas_init_label, cfg["tco_saas_initial"], YEN_FMT),
    ]
    rows_by_label: dict[str, int] = {}
    for i, (label, val, fmt) in enumerate(presets):
        r = 7 + i
        rows_by_label[label] = r
        _cell(ws, r, 2, label, font=BODY_BOLD, fill=FILL_SUBHEAD, align=LEFT)
        _cell(ws, r, 3, val, font=INPUT_FONT, fill=FILL_INPUT, align=RIGHT, fmt=fmt)

    def C(label: str) -> str:
        return f"$C${rows_by_label[label]}"

    _heading(ws, 17, 2, "5 年総コスト比較", font=H2)

    headers = ("形態", "初期費用", "年間運用費 × N", "5 年総額")
    for j, h in enumerate(headers):
        _cell(ws, 18, 2 + j, h, font=HEAD_FONT, fill=FILL_HEAD, align=CENTER)
    ws.row_dimensions[18].height = 26

    # 中堅 IT
    _cell(ws, 19, 2, "自社開発 (中堅 IT 会社)", font=BODY_BOLD, align=LEFT)
    _cell(ws, 19, 3, f"={C('中堅 IT 開発初期費 (¥)')}", font=BODY, align=RIGHT, fmt=YEN_FMT)
    _cell(ws, 19, 4, f"={C('中堅 IT 年間運用費 (¥)')}*{C('運用年数')}", font=BODY, align=RIGHT, fmt=YEN_FMT)
    _cell(ws, 19, 5, "=C19+D19", font=BODY_BOLD, fill=FILL_DANGER, align=RIGHT, fmt=YEN_FMT)

    # 大手 SI
    _cell(ws, 20, 2, "自社開発 (大手 SIer)", font=BODY_BOLD, align=LEFT)
    _cell(ws, 20, 3, f"={C('大手 SIer 開発初期費 (¥)')}", font=BODY, align=RIGHT, fmt=YEN_FMT)
    _cell(ws, 20, 4, f"={C('大手 SIer 年間運用費 (¥)')}*{C('運用年数')}", font=BODY, align=RIGHT, fmt=YEN_FMT)
    _cell(ws, 20, 5, "=C20+D20", font=BODY_BOLD, fill=FILL_DANGER, align=RIGHT, fmt=YEN_FMT)

    # SaaS
    _cell(ws, 21, 2, f"{cfg['service_name']} SaaS (Enterprise)", font=BODY_BOLD, fill=FILL_OUTPUT, align=LEFT)
    _cell(ws, 21, 3, f"={C(saas_init_label)}", font=BODY, fill=FILL_OUTPUT, align=RIGHT, fmt=YEN_FMT)
    monthly_total_expr = (
        f"({C(cfg['service_name'] + ' Enterprise 月額/店舗 (¥)')}"
        f"*{C('店舗数')}+{C(cfg['service_name'] + ' Enterprise 本部/月 (¥)')})"
    )
    _cell(ws, 21, 4, f"={monthly_total_expr}*12*{C('運用年数')}",
          font=BODY, fill=FILL_OUTPUT, align=RIGHT, fmt=YEN_FMT)
    _cell(ws, 21, 5, "=C21+D21",
          font=Font(name="Yu Gothic", size=12, bold=True, color="047857"),
          fill=FILL_OUTPUT, align=RIGHT, fmt=YEN_FMT)

    # 倍率
    _heading(ws, 23, 2, "コスト比較", font=H2)
    _cell(ws, 24, 2, f"中堅 IT 開発 / {cfg['service_name']} SaaS", font=BODY_BOLD, align=LEFT)
    _cell(ws, 24, 3, "=E19/E21", font=BODY_BOLD, fill=FILL_OUTPUT, align=RIGHT, fmt='0.00"倍"')
    _cell(ws, 25, 2, f"大手 SIer 開発 / {cfg['service_name']} SaaS", font=BODY_BOLD, align=LEFT)
    _cell(ws, 25, 3, "=E20/E21", font=BODY_BOLD, fill=FILL_OUTPUT, align=RIGHT, fmt='0.00"倍"')

    _cell(ws, 27, 2,
          "※ 中堅 IT 開発の年間運用費は、サーバ運用・保守エンジニア 1〜2 名分を想定。\n"
          "※ 大手 SIer は運用も SLA 付き保守契約での想定。\n"
          "※ SaaS の年間運用費には、機能改善・セキュリティアップデート・サポート全て含む。",
          font=SMALL, align=LEFT, border=None)
    ws.merge_cells("B27:E29")
    for r in (27, 28, 29):
        ws.row_dimensions[r].height = 16


# ============================================================
# 5. オプション
# ============================================================

def build_options(wb: Workbook, cfg: dict) -> None:
    ws = wb.create_sheet("オプション")
    _set_widths(ws, [4, 44, 28, 38, 4])

    _heading(ws, 2, 2, "オプション・アドオン (全プラン共通)", font=H1)

    headers = ("項目", "料金 (税抜)", "備考")
    for j, h in enumerate(headers):
        _cell(ws, 4, 2 + j, h, font=HEAD_FONT, fill=FILL_HEAD, align=CENTER)
    ws.row_dimensions[4].height = 26

    options = [
        ("標準 6 言語以外の言語追加 (タイ語・タガログ語等)",
         "初期 ¥100,000 + ¥3,000/月", "翻訳精度評価込み"),
        ("カスタム機能開発",
         "¥500,000〜", "規模・要件に応じ個別見積"),
        ("POS / 勤怠管理 / シフト管理連携",
         "¥300,000〜¥2,000,000", "対象システムにより変動"),
        ("専用ロゴ・ブランド統合",
         "¥150,000 (一括)", "管理者・スタッフ画面"),
        ("オンサイトサポート (店舗訪問)",
         "¥150,000/日", "交通費別"),
        ("プレミアム SLA (24 時間緊急対応)",
         "¥150,000/月", "Enterprise の標準オプション"),
        ("月次定例ミーティング・改善レポート",
         "¥50,000/月", "Enterprise 既定"),
        ("データ保持期間延長 (60 日 → 6 ヶ月)",
         "¥10,000/月", ""),
        ("データ保持期間延長 (60 日 → 12 ヶ月)",
         "¥20,000/月", ""),
        ("ペネトレーションテスト (年 1 回)",
         "¥500,000〜¥1,500,000/年", "Enterprise 既定"),
        ("DPIA (データ保護影響評価) 個別実施",
         "¥300,000〜", ""),
        ("管理者向けスタッフ教育コンテンツ作成",
         "¥200,000〜", "業種別カリキュラム作成"),
    ]

    for i, (item, price, note) in enumerate(options):
        r = 5 + i
        fill = FILL_ALT if i % 2 == 0 else None
        _cell(ws, r, 2, item, font=BODY, fill=fill, align=LEFT)
        _cell(ws, r, 3, price, font=BODY_BOLD, fill=fill, align=CENTER)
        _cell(ws, r, 4, note, font=SMALL, fill=fill, align=LEFT)
        ws.row_dimensions[r].height = 22

    end_row = 5 + len(options) + 2
    _cell(ws, end_row, 2,
          "※ 上記オプションは事前見積もりとお客様の承諾の上で実施いたします。\n"
          "※ 既存契約の途中追加も可能です。次回請求月から適用されます。",
          font=SMALL, align=LEFT, border=None)
    ws.merge_cells(start_row=end_row, start_column=2, end_row=end_row, end_column=4)


# ============================================================
# Main
# ============================================================

def build_workbook(cfg: dict) -> Workbook:
    wb = Workbook()
    build_cover(wb, cfg)
    build_calculator(wb, cfg)
    build_plan_comparison(wb, cfg)
    build_tco(wb, cfg)
    build_options(wb, cfg)
    return wb


def main():
    out_dir = Path(__file__).resolve().parent
    for variant_key, cfg in CONFIGS.items():
        wb = build_workbook(cfg)
        out_path = out_dir / cfg["output_filename"]
        wb.save(out_path)
        print(f"saved: {out_path.name}  ({cfg['service_name']})")


if __name__ == "__main__":
    main()
